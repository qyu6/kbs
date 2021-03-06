


from os.path import exists
import openpyxl
import os
import pandas as pd
import re
from collections import Counter
import streamlit as st
from PIL import Image

# Searching engine main function:
def main():

    pd.set_option('display.max_colwidth',None)

    result = 'searchoutput.csv'
    if exists(result):
        os.remove(result)
    # 创建结果文件
    wbResult = openpyxl.Workbook()
    wsResult = wbResult.worksheets[0]
    wsResult.append(['result'])
    # 读取原表两次，一次用来进行建表输入，一次用来做对应的输入
    wb = openpyxl.load_workbook('SourceDB.xlsx')
    input_excel = 'SourceDB.xlsx'
    data = pd.read_excel(input_excel)
    ws = wb.worksheets[0]
    # 原表空白部分用*填充
    for k in range(1,ws.max_column+1):
        for i in range(1,ws.max_row+1):
            if ws.cell(row=i,column=k).value is None:
                ws.cell(i,k,'****')


    # input_word = input("请输入搜索内容:").strip().lower()
    # st.subheader('🐼[T.Q Knowledge Base]')
    input_word1 = st.text_input('©🐼 | Last release:2022/3/3 | Search your knowledge from here...','')
    input_word = input_word1.strip().lower()
    input_word_exist = re.sub(u"([u4e00-\u9fa5\u0030-\u0039\u0041-\u005a\u0061-\u007a])","",input_word)
    input_word = input_word.split()



    result_list = []
    for index,row in enumerate(ws.rows):

        if index == 0:
            continue
        rs_list = list(map(lambda cell: cell.value, row))
        list_str = "".join('%s' %id for id in rs_list).replace("\n"," ").replace("\n"," ").replace("\t"," ").replace("\r"," ").lower()
        result_list.append([index, list_str])



    def search_onebyone(input_word_exist, input_word_list, result_list):
        new_list = []
        dict_list = []
        new_list_count = []
        # 精确匹配
        for i in range(len(result_list)):
            for m in input_word_list:
                pattern = m
                regex = re.compile(pattern)
                nz = regex.search(result_list[i][1])
                if nz:
                    new_list.append([len(nz.group()),nz.start(),result_list[i][0]-1])
                    new_list_count.append(result_list[i][0]-1)

        new_list = sorted(new_list)
        new_index = [x for _,_,x in new_list]
        new_index = sorted(set(new_index),key=new_index.index)

        # 计数，只有当输入的全部单词全部出现以后，才取出
        dict_list.append([k for k,v in Counter(new_list_count).items() if v == len(input_word_list)])
        for m in dict_list:
            result_index = m
        temp = [j for j in new_index if j in result_index]
        return temp
    result = search_onebyone(input_word_exist, input_word, result_list)



    def display_highlighted_words(df, keywords):
        head = """
        <talbe>
            <thead>
                """ + \
                    "".join(["<th> %s </th>" % c for c in df.columns])\
                    + """
            </thead>
        </table>"""

        head = """
        <table>
            <thead>
                <th> Keywords </th><th> Content </th>
            </thead>
        </table>
        """

        for i,r in df.iterrows():
            row = "<tr>"
            for c in df.columns:
                matches = []
                for k in keywords:
                    for match in re.finditer(k, str(r[c])):
                        matches.append(match)

                # reverse sorting
                matches = sorted(matches, key = lambda x: x.start(), reverse=True)

                # building HTML row
                cell = str(r[c])

                # print(cell)
                for match in matches:
                    cell = cell[:match.start()] +\
                        "<span style='color:red;background-color:yellow'> %s </span>" % cell[match.start():match.end()] +\
                        cell[match.end():]

                row += "<td> %s </td>" % cell

            row += "</tr>"

            head += row


            head += "</tbody></table>"

            return head

    # htmlcode1 = display_highlighted_words(dftest, input_word)
    # st.markdown(htmlcode1, unsafe_allow_html=True)


    if len(input_word)>0:
        st.table(data.loc[(x for x in result)])
        # dftest = data.loc[(x for x in result)]
        # htmlcode1 = display_highlighted_words(dftest, input_word)
        # st.markdown(htmlcode1, unsafe_allow_html=True)

    # - ↓ Incase need to save search result files.
    # data.loc[(x for x in result)].to_csv('searchoutput.csv', encoding= 'utf_8_sig')



from SessionState import get
session_state = get(AccessCode='')


if session_state.AccessCode != '211':

    head_placeholder = st.empty()
    pwd_placeholder = st.empty()

    head = head_placeholder.subheader('🔒')
    pwd = pwd_placeholder.text_input("Access Code", value="", type="password")
    
    session_state.AccessCode = pwd
    if session_state.AccessCode == '211':
        head_placeholder.empty()
        pwd_placeholder.empty()
        main()
        # st.sidebar.write('Test passed')

    elif session_state.AccessCode != '':
        st.error("Access denied! Please scan below **[QR-Code]** below to get access code.")
        qrimage = Image.open('pics/QR-code.jpg')
        st.image(qrimage)
    else:
        nothing = 'do nothing'
else:
    main()
    # st.write('Test passed')